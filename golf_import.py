#test program to modify a cell on a spreadsheet

import openpyxl

#loads sheet names
wb = openpyxl.load_workbook('golf.xlsx')

#prints sheet names
#print(wb.sheetnames)

#creates new sheet
#wb.create_sheet()

#print new list of sheetnames
#print(wb.sheetnames)

sheet = wb['Sheet1']

sheet['A1'] = 'Goodbye World'

tester = sheet['A1'].value

print(tester)

wb.save('golf_import_test_save.xlsx')
